import argparse
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

TEMPLATE_HEADERS = [
    "商品ID",
    "创建时间",
    "商品标题*",
    "商品属性*",
    "商品类型",
    "商品副标题",
    "商品描述",
    "简短描述",
    "内部商品名",
    "SEO标题",
    "SEO描述",
    "SEO URL Handle",
    "SEO关键词",
    "商品上架",
    "商品收税",
    "商品spu",
    "虚拟销量值",
    "跟踪库存",
    "库存规则*",
    "专辑名称",
    "标签",
    "供应商名称",
    "款式1",
    "款式2",
    "款式3",
    "商品售价*",
    "商品原价",
    "商品SKU",
    "商品重量(kg)",
    "商品条形码",
    "商品库存",
    "变体备注1",
    "变体备注2",
    "商品图片*",
]


@dataclass
class BaseProject:
    title: str
    short_desc: str
    progress_pct: Optional[int]
    progress_text: str
    link: str


@dataclass
class DetailData:
    url: str
    title: str
    short_desc: str
    story_html: str
    seo_title: str
    seo_description: str
    seo_keywords: str
    canonical_url: str
    image_url: str


@dataclass
class RewardData:
    title: str
    price_text: str
    image_url: str


def load_base_projects(path: str, min_progress: int, limit: Optional[int]) -> List[BaseProject]:
    wb = load_workbook(path)
    ws = wb.active
    projects: List[BaseProject] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[4]:
            continue
        progress = row[2] or 0
        if progress is None or progress < min_progress:
            continue
        projects.append(BaseProject(row[0] or "", row[1] or "", progress, row[3] or "", row[4].strip()))
        if limit and len(projects) >= limit:
            break
    logger.info(f"Loaded {len(projects)} base projects from {path}")
    return projects


def parse_detail_page(url: str, html: str, fallback_desc: str) -> Tuple[DetailData, List[RewardData]]:
    soup = BeautifulSoup(html, "lxml")

    def _get(selector: str) -> str:
        el = soup.select_one(selector)
        return el.get("content", "").strip() if el and el.has_attr("content") else (el.get_text(" ", strip=True) if el else "")

    h1 = soup.select_one("h1")
    title = h1.get_text(" ", strip=True) if h1 else _get("meta[property='og:title']")
    seo_title = _get("meta[property='og:title']") or (soup.title.string.strip() if soup.title else title)
    seo_description = _get("meta[name='description']") or _get("meta[property='og:description']")
    seo_keywords = _get("meta[name='keywords']")
    canonical_el = soup.select_one("link[rel='canonical']")
    canonical_url = canonical_el.get("href", "").strip() if canonical_el else url
    image_url = _get("meta[property='og:image']")
    short_desc = seo_description or fallback_desc

    story_container = soup.select_one("#story + .story-content .rte__content") or soup.select_one(".story-content .rte__content") or soup.select_one(".story-content")
    story_html = story_container.decode_contents() if story_container else short_desc

    detail = DetailData(
        url=url,
        title=title or fallback_desc,
        short_desc=short_desc,
        story_html=story_html,
        seo_title=seo_title,
        seo_description=seo_description,
        seo_keywords=seo_keywords,
        canonical_url=canonical_url,
        image_url=image_url,
    )
    rewards = parse_rewards_from_soup(soup)
    return detail, rewards


def parse_rewards_from_soup(soup: BeautifulSoup) -> List[RewardData]:
    cards = soup.select("[data-test-id^='reward-']")
    rewards: List[RewardData] = []
    for card in cards:
        header = card.find("header")
        title = header.h3.get_text(" ", strip=True) if header and header.h3 else ""
        price_el = header.find("p", class_="support-700") if header else None
        price_text = price_el.get_text(" ", strip=True) if price_el else ""
        img_el = card.select_one("img")
        image_url = img_el.get("src", "") if img_el else ""
        rewards.append(RewardData(title=title, price_text=price_text, image_url=image_url))
    return rewards


def normalize_price(text: str) -> Tuple[Optional[str], Optional[float]]:
    if not text:
        return None, None
    parts = text.strip().split()
    currency = None
    if parts and not re.match(r"^[\d,\.]+$", parts[0]):
        currency = parts[0]
    match = re.search(r"[\d,.]+", text)
    if not match:
        return currency, None
    value = float(match.group(0).replace(",", ""))
    return currency, value


def collect_detail_and_rewards(projects: List[BaseProject]) -> Tuple[Dict[str, DetailData], Dict[str, List[RewardData]]]:
    details: Dict[str, DetailData] = {}
    rewards: Dict[str, List[RewardData]] = {}

    def fetch_html(page, url: str) -> str:
        logger.debug(f"Requesting {url}")
        page.goto(url, wait_until="domcontentloaded", timeout=120000)
        page.wait_for_timeout(6000)
        html = page.content()
        title = page.title()
        if "Just a moment" not in (title or "") and "cf-browser-verification" not in html:
            return html
        last_html = html
        for attempt in range(3):
            wait_ms = 6000 + attempt * 4000
            logger.debug(f"Waiting {wait_ms/1000:.1f}s for Cloudflare challenge on {url}")
            page.wait_for_timeout(wait_ms)
            html = page.content()
            title = page.title()
            last_html = html
            if "Just a moment" not in (title or "") and "cf-browser-verification" not in html:
                return html
        logger.warning(f"Cloudflare challenge persisted for {url}")
        return last_html

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1280, "height": 2200})
        for idx, project in enumerate(projects, 1):
            page = context.new_page()
            logger.info(f"[{idx}/{len(projects)}] Fetching detail {project.link}")
            detail_html = fetch_html(page, project.link)
            detail_data, parsed_rewards = parse_detail_page(project.link, detail_html, project.short_desc)
            details[project.link] = detail_data
            logger.info(f"Parsed {len(parsed_rewards)} rewards for {project.link}")
            rewards[project.link] = parsed_rewards
            page.close()
        browser.close()
    return details, rewards


def slug_from_url(url: str) -> str:
    parsed = urlparse(url)
    return parsed.path.strip("/").replace("/", "-")


def slugify(text: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9-]+", "-", text)
    slug = re.sub("-+", "-", slug).strip("-")
    return slug or "item"


def build_rows(
    projects: List[BaseProject],
    details: Dict[str, DetailData],
    rewards: Dict[str, List[RewardData]],
    max_rewards: Optional[int] = None,
) -> List[List[str]]:
    rows: List[List[str]] = []
    for project in projects:
        detail = details.get(project.link)
        reward_list = rewards.get(project.link, [])
        project_title = (detail.title if detail else project.title).strip()
        short_desc = (detail.short_desc if detail else project.short_desc).strip()
        long_desc = detail.story_html if detail else short_desc
        seo_title = detail.seo_title if detail else project_title
        seo_desc = detail.seo_description if detail else short_desc
        seo_keywords = detail.seo_keywords if detail else ""
        canonical_url = detail.canonical_url if detail else project.link
        handle = slug_from_url(canonical_url)
        sku_base = handle
        main_image = detail.image_url if detail else ""
        category_name = "Gaming Hardware"

        base_row = [""] * len(TEMPLATE_HEADERS)
        base_row[2] = project_title
        base_row[3] = "M"
        base_row[4] = category_name
        base_row[5] = short_desc
        base_row[6] = long_desc
        base_row[7] = short_desc
        base_row[8] = project_title
        base_row[9] = seo_title
        base_row[10] = seo_desc
        base_row[11] = handle
        base_row[12] = seo_keywords
        base_row[13] = "Y"
        base_row[14] = "N"
        base_row[15] = sku_base
        base_row[17] = "N"
        base_row[18] = 2
        base_row[20] = "Kickstarter, Category270"
        base_row[21] = "Kickstarter"
        base_row[22] = "Type"
        base_row[27] = sku_base
        base_row[30] = "999"
        base_row[31] = ""
        base_row[32] = ""
        base_row[33] = main_image
        rows.append(base_row)

        reward_iterable = reward_list
        if max_rewards is not None:
            reward_iterable = reward_list[:max_rewards]
        if not reward_iterable:
            reward_iterable = [
                RewardData(title=project_title, price_text="", image_url=main_image)
            ]

        for idx, reward in enumerate(reward_iterable, 1):
            row = [""] * len(TEMPLATE_HEADERS)
            currency, price_value = normalize_price(reward.price_text)
            display_price = reward.price_text if reward.price_text else (f"{price_value:.2f}" if price_value else "")
            row[2] = project_title
            row[3] = "P"
            row[4] = category_name
            row[5] = short_desc
            row[6] = long_desc
            row[7] = short_desc
            row[8] = project_title
            row[9] = seo_title
            row[10] = seo_desc
            row[11] = handle
            row[12] = seo_keywords
            row[13] = "Y"
            row[14] = "N"
            variant_slug = slugify(reward.title or f"variant-{idx}")[:30] or f"variant-{idx}"
            base_part = sku_base
            max_len = 63
            combined = f"{base_part}-{variant_slug}"
            if len(combined) > max_len:
                available = max_len - len(variant_slug) - 1
                if available < 1:
                    trimmed_variant = variant_slug[: max_len - 2]
                    if not trimmed_variant:
                        trimmed_variant = "var"
                    variant_slug = trimmed_variant
                    available = max_len - len(variant_slug) - 1
                base_part = base_part[: max(1, available)]
                combined = f"{base_part}-{variant_slug}"
            sku = combined[:max_len]
            row[15] = sku
            row[17] = "N"
            row[18] = 2
            row[20] = "Kickstarter, Category270"
            row[21] = "Kickstarter"
            row[22] = reward.title or f"Variant {idx}"
            row[25] = display_price
            row[26] = display_price
            row[27] = sku
            row[30] = "999"
            row[31] = ""
            row[32] = ""
            row[33] = reward.image_url or main_image
            rows.append(row)
    return rows


def write_to_template(rows: List[List[str]], template_path: str) -> None:
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    wb = load_workbook(template_path)
    ws = wb.active
    if ws.max_row < 2:
        ws.insert_rows(2)
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.cell(row=2, column=col_idx, value=None)
    start_row = 3
    for idx, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + idx, column=col_idx, value=value)
    wb.save(template_path)
    logger.info(f"Wrote {len(rows)} rows to {template_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build ks.xlsx from Kickstarter data")
    parser.add_argument("--list-file", default=os.path.join("output", "category_270_raised2_all.xlsx"))
    parser.add_argument("--template", default=os.path.join("output", "ks.xlsx"))
    parser.add_argument("--min-progress", type=int, default=100)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--max-rewards", type=int, help="Max rewards per project to export")
    args = parser.parse_args()

    projects = load_base_projects(args.list_file, args.min_progress, args.limit)
    if not projects:
        logger.error("No projects to process.")
        return
    details, rewards = collect_detail_and_rewards(projects)
    rows = build_rows(projects, details, rewards, max_rewards=args.max_rewards)
    write_to_template(rows, args.template)


if __name__ == "__main__":
    main()
