import os
import sys
from dataclasses import dataclass
from typing import List, Optional, Set

from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


@dataclass
class Detail:
    url: str
    title: str
    short_desc_en: str
    seo_title: str
    seo_description: str
    seo_keywords: str
    canonical_url: str


def load_links_from_xlsx(path: str) -> List[str]:
    wb = load_workbook(path)
    ws = wb.active
    links: List[str] = []
    seen: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        link = (row[4] or "").strip()  # column "link"
        if link and link not in seen:
            seen.add(link)
            links.append(link)
    return links


def parse_detail(url: str, html: str) -> Detail:
    soup = BeautifulSoup(html, "lxml")
    title = soup.select_one("meta[property='og:title']")
    meta_desc = soup.select_one("meta[name='description']") or soup.select_one("meta[property='og:description']")
    keywords = soup.select_one("meta[name='keywords']")
    canonical = soup.select_one("link[rel='canonical']")
    h1 = soup.select_one("h2[itemprop='name'], h1")
    short_blurb = soup.select_one("meta[name='description']") or soup.select_one("meta[property='og:description']")

    return Detail(
        url=url,
        title=(h1.get_text(" ", strip=True) if h1 else (title.get("content", "") if title else "")),
        short_desc_en=(short_blurb.get("content", "") if short_blurb else ""),
        seo_title=(title.get("content", "") if title else (soup.title.string.strip() if soup.title else "")),
        seo_description=(meta_desc.get("content", "") if meta_desc else ""),
        seo_keywords=(keywords.get("content", "") if keywords else ""),
        canonical_url=(canonical.get("href", "") if canonical else url),
    )


def scrape_details(links: List[str]) -> List[Detail]:
    results: List[Detail] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1280, "height": 2000})
        for idx, link in enumerate(links, 1):
            page = context.new_page()
            logger.info(f"[{idx}/{len(links)}] Fetching {link}")
            page.goto(link, wait_until="domcontentloaded", timeout=120000)
            page.wait_for_timeout(5000)
            html = page.content()
            results.append(parse_detail(link, html))
            page.close()
        browser.close()
    return results


def save_details(items: List[Detail], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "details"
    ws.append(["url", "title", "short_desc_en", "seo_title", "seo_description", "seo_keywords", "canonical_url"])
    for item in items:
        ws.append(
            [
                item.url,
                item.title,
                item.short_desc_en,
                item.seo_title,
                item.seo_description,
                item.seo_keywords,
                item.canonical_url,
            ]
        )
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    logger.info(f"Saved {len(items)} rows to {filepath}")


def main():
    if len(sys.argv) < 2:
        logger.error("Usage: python detail_pages.py <list_xlsx_path>")
        sys.exit(1)
    list_path = sys.argv[1]
    links = load_links_from_xlsx(list_path)
    logger.info(f"Loaded {len(links)} unique links from {list_path}")
    details = scrape_details(links)
    output_path = os.path.join(os.path.dirname(list_path), "category_detail_pages.xlsx")
    save_details(details, output_path)


if __name__ == "__main__":
    main()
