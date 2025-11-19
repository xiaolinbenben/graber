import os
import sys
from dataclasses import dataclass
from typing import List, Optional, Set

from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


@dataclass
class Reward:
    url: str
    pledge_amount: str
    title: str
    description: str
    ships_to: str
    delivery: str
    backer_limit: str


def load_links_from_xlsx(path: str) -> List[str]:
    wb = load_workbook(path)
    ws = wb.active
    links: List[str] = []
    seen: Set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        link = (row[4] or "").strip()  # column "link"
        if link and link not in seen:
            seen.add(link)
            links.append(link.rstrip("/"))
    return links


def parse_rewards(url: str, html: str) -> List[Reward]:
    soup = BeautifulSoup(html, "lxml")
    rewards: List[Reward] = []

    # Reward cards appear under data-test-id attributes; selectors may vary, keep it forgiving.
    cards = soup.select("[data-test-id='reward-card'], div.reward, li.reward, div[data-reward-id]")
    for card in cards:
        pledge = ""
        amt_el = card.select_one("[data-test-id='reward-amount'], .pledge__amount, .pledge__currency")
        if amt_el:
            pledge = amt_el.get_text(" ", strip=True)

        title_el = card.select_one("[data-test-id='reward-title'], .pledge__title, h3")
        desc_el = card.select_one("[data-test-id='reward-description'], .pledge__reward-description, .desc, p")
        ship_el = card.find(string=lambda s: isinstance(s, str) and "Ships to" in s)
        delivery_el = card.find(string=lambda s: isinstance(s, str) and ("Delivery" in s or "Estimate" in s))
        limit_el = card.find(string=lambda s: isinstance(s, str) and ("backer" in s or "Backer" in s))

        rewards.append(
            Reward(
                url=url,
                pledge_amount=pledge,
                title=title_el.get_text(" ", strip=True) if title_el else "",
                description=desc_el.get_text(" ", strip=True) if desc_el else "",
                ships_to=ship_el.strip() if ship_el else "",
                delivery=delivery_el.strip() if delivery_el else "",
                backer_limit=limit_el.strip() if limit_el else "",
            )
        )
    return rewards


def scrape_rewards(links: List[str]) -> List[Reward]:
    results: List[Reward] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1280, "height": 2000})
        for idx, link in enumerate(links, 1):
            rewards_url = f"{link}/rewards"
            page = context.new_page()
            logger.info(f"[{idx}/{len(links)}] Fetching rewards {rewards_url}")
            page.goto(rewards_url, wait_until="domcontentloaded", timeout=120000)
            page.wait_for_timeout(5000)
            html = page.content()
            rewards = parse_rewards(rewards_url, html)
            results.extend(rewards)
            page.close()
        browser.close()
    return results


def save_rewards(items: List[Reward], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "rewards"
    ws.append(["url", "pledge_amount", "title", "description", "ships_to", "delivery", "backer_limit"])
    for item in items:
        ws.append(
            [
                item.url,
                item.pledge_amount,
                item.title,
                item.description,
                item.ships_to,
                item.delivery,
                item.backer_limit,
            ]
        )
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    logger.info(f"Saved {len(items)} reward rows to {filepath}")


def main():
    if len(sys.argv) < 2:
        logger.error("Usage: python rewards_pages.py <list_xlsx_path>")
        sys.exit(1)
    list_path = sys.argv[1]
    links = load_links_from_xlsx(list_path)
    logger.info(f"Loaded {len(links)} unique links from {list_path}")
    rewards = scrape_rewards(links)
    output_path = os.path.join(os.path.dirname(list_path), "category_reward_pages.xlsx")
    save_rewards(rewards, output_path)


if __name__ == "__main__":
    main()
